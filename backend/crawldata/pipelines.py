# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import datetime
import pandas as pd
from openpyxl import load_workbook
import os

class CrawldataPipeline(object):
    def process_item(self, item, spider):
        return item

class NationalPMPipeline(object):
    national_pm = None
    def open_spider(self, spider):
        self.national_pm = pd.DataFrame(columns=('城市', '监测点', 'AQI', '空气质量指数类别', '首要污染物', 'PM2.5细颗粒物（μg/m3)',
                      'PM10可吸入颗粒物（μg/m3）', 'CO一氧化碳（mg/m3）', 'NO2二氧化氮（μg/m3）',
                      'O3臭氧1小时平均（μg/m3）', 'O3_8h臭氧8小时平均（μg/m3）', 'SO2二氧化硫（μg/m3）',
                      '数据更新时间'),index=None)

    def process_item(self, item, spider):
        self.national_pm.loc[self.national_pm.shape[0]] = [i for i in item.values()]
        return item

    def close_spider(self, spider):
        date = datetime.date.today().strftime('%y%m%d')
        path = os.path.join("static/" + '国内空气污染实时数据.xlsx')
        p1 = os.path.exists(path)
        if p1:
            writer = pd.ExcelWriter(path, engine='openpyxl')
            book = load_workbook(writer.path)
            writer.book = book
            self.national_pm.to_excel(writer, '空气污染物质1-国内', index=None)
        else:
            writer = pd.ExcelWriter(path)
            self.national_pm.to_excel(writer, '空气污染物质1-国内', index=None)
        writer.save()
        writer.close()

class WorldPMPipeline(object):
    world_pm = None
    def open_spider(self, spider):
        self.world_pm = pd.DataFrame(columns=('时间', '监测点', 'AQI', '空气质量指数类别', '首要污染物', 'PM2.5',
                      'PM10', 'O3', 'NO2', 'SO2', 'CO', '温度','空气压力','湿度','风'),index=None)

    def process_item(self, item, spider):
        self.world_pm.loc[self.world_pm.shape[0]] = [i for i in item.values()]
        return item

    def close_spider(self, spider):
        date = datetime.date.today().strftime('%y%m%d')
        path = os.path.join("static/" + '世界空气污染实时数据.xlsx')
        p1 = os.path.exists(path)
        if p1:
            writer = pd.ExcelWriter(path, engine='openpyxl')
            book = load_workbook(writer.path)
            writer.book = book
            self.world_pm.to_excel(writer, '空气污染物质2-世界', index=None)
        else:
            writer = pd.ExcelWriter(path)
            self.world_pm.to_excel(writer, '空气污染物质2-世界', index=None)
        writer.save()
        writer.close()

class NationalWaterPollutionPipeline(object):
    national_waterpollution = None
    def open_spider(self, spider):
        self.national_waterpollution = pd.DataFrame(columns=('断面名称', '测量时间', 'pH', '溶解氧', '氨氮', '高锰酸盐指数',
                      '总有机碳','水质类别','断面属性','站点情况'),index=None)

    def process_item(self, item, spider):
        self.national_waterpollution.loc[self.national_waterpollution.shape[0]] = [i for i in item.values()]
        return item

    def close_spider(self, spider):
        date = datetime.date.today().strftime('%y%m%d')
        path = os.path.join("static/" + '国内水体污染实时数据.xlsx')
        p1 = os.path.exists(path)
        if p1:
            writer = pd.ExcelWriter(path, engine='openpyxl')
            book = load_workbook(writer.path)
            writer.book = book
            self.national_waterpollution.to_excel(writer,'水体污染物质-全国',index=None)
        else:
            writer = pd.ExcelWriter(path)
            self.national_waterpollution.to_excel(writer, '水体污染物质-全国', index=None)
        writer.save()
        writer.close()

class NationalSolidPollutionPipeline(object):
    national_solidpollution = None
    def open_spider(self, spider):
        self.national_solidpollution = pd.DataFrame(columns=('年份', '地区', '指标', '数值', '单位', '来源'),index=None)

    def process_item(self, item, spider):
        self.national_solidpollution.loc[self.national_solidpollution.shape[0]] = [i for i in item.values()]
        return item

    def close_spider(self, spider):
        date = datetime.date.today().strftime('%y%m%d')
        path = os.path.join("static/" + '国内固体废物实时数据.xlsx')
        p1 = os.path.exists(path)
        if p1:
            writer = pd.ExcelWriter(path,engine='openpyxl')
            book = load_workbook(writer.path)
            writer.book = book
            self.national_solidpollution.to_excel(writer,'固体废弃物-全国',index=None)
        else:
            writer = pd.ExcelWriter(path)
            self.national_solidpollution.to_excel(writer, '固体废弃物-全国', index=None)
        writer.save()
        writer.close()

# class NationalSolidPollutionPipeline(object):
#     national_solidpollution = None
#     def open_spider(self, spider):
#         self.national_solidpollution = pd.DataFrame(columns=('年份', '地区', '指标', '数值', '单位', '来源'),index=None)
#
#     def process_item(self, item, spider):
#         self.national_solidpollution.loc[self.national_solidpollution.shape[0]] = [i for i in item.values()]
#         return item
#
#     def close_spider(self, spider):
#         date = datetime.date.today().strftime('%y%m%d')
#         path = os.path.join(os.getcwd() + "\\" +date+'废弃物实时数据.xlsx')
#         p1 = os.path.exists(path)
#         if p1:
#             writer = pd.ExcelWriter(date+'废弃物实时数据.xlsx',engine='openpyxl')
#             book = load_workbook(writer.path)
#             writer.book = book
#             self.national_solidpollution.to_excel(writer,'固体废弃物-全国',index=None)
#         else:
#             writer = pd.ExcelWriter(date + '废弃物实时数据.xlsx')
#             self.national_solidpollution.to_excel(writer, '固体废弃物-全国', index=None)
#         writer.save()
#         writer.close()
